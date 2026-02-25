import streamlit as st
import pandas as pd
import openpyxl
import os
import subprocess
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# --- CONFIGURAÇÕES DE SERVIDOR DE E-MAIL (VIA APPIA) ---
SMTP_SERVER = "10.254.255.25"
SMTP_PORT = 25

def enviar_email_corporativo(remetente_email, destinatarios, assunto, corpo_html, reply_to, anexos=None):
    if anexos is None:
        anexos = []

    msg = MIMEMultipart()
    msg["From"] = remetente_email
    msg["To"] = ", ".join(destinatarios)
    msg["Subject"] = assunto
    if reply_to:
        msg["Reply-To"] = reply_to

    # Corpo em HTML
    msg.attach(MIMEText(corpo_html, "html"))

    # Anexos
    for caminho_arquivo in anexos:
        if not os.path.exists(caminho_arquivo):
            continue
        with open(caminho_arquivo, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        nome_arquivo = os.path.basename(caminho_arquivo)
        part.add_header("Content-Disposition", f'attachment; filename="{nome_arquivo}"')
        msg.attach(part)

    # Envia via Relay Interno da Via Appia
    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.sendmail(remetente_email, destinatarios, msg.as_string())
    server.quit()

# --- CONFIGURAÇÕES DO APP STREAMLIT ---
TEMPLATE_PATH = "FORMULARIO FATURAS FI NOVO v1.xlsx" 

st.set_page_config(page_title="Sistema de Reembolsos - Via Appia", layout="wide")

st.title("🧾 Portal de Relatório de Reembolsos")
st.write("Preencha os dados abaixo e anexe seus comprovantes. Após processamento, o formulário e os comprovantes serão enviados à esteira fiscal.")

col1, col2 = st.columns(2)
with col1:
    st.subheader("Dados do Solicitante")
    solicitante = st.text_input("Solicitante (Responsável)")
    solicitante_email = st.text_input("Seu e-mail (para receber a cópia/contato)", placeholder="nome@viaappia.com.br")
with col2:
    st.subheader("Dados do Colaborador a Reembolsar")
    colaborador_nome = st.text_input("Nome do Colaborador")
    colaborador_cpf = st.text_input("CPF")
    nivel_hierarquico = st.text_input("Nível Hierárquico")
    fornecedor_sap = st.text_input("Nº Fornecedor SAP")

st.divider()
st.subheader("🛒 Despesas Gerais")
st.write("Insira as despesas convencionais (Hospedagem, Alimentação, etc.):")
df_despesas = pd.DataFrame(columns=["Data (DD/MM/AAA)", "Conta Razão", "Centro de Custo", "Motivo ou Justificativa", "Qtde", "Valor Gasto (R$)"])
despesas_editadas = st.data_editor(df_despesas, num_rows="dynamic", width='stretch')

st.divider()
st.subheader("🚗 Reembolso de Quilometragem")
df_km = pd.DataFrame(columns=["Data (DD/MM/AAA)", "Conta Razão", "Centro de Custo", "Motivo/Origem>Destino", "Km (Qtde)", "Valor/Km (R$)", "Valor Gasto (R$)"])
km_editados = st.data_editor(df_km, num_rows="dynamic", width='stretch')

# --- NOVA SESSÃO: COMPROVANTES ---
st.divider()
st.subheader("📎 Anexar Comprovantes")
st.write("Suba os arquivos das notas fiscais e recibos. Eles serão enviados em anexo junto do formulário.")
comprovantes_upload = st.file_uploader(
    "Arraste ou selecione seus comprovantes (PDF, JPG, PNG)", 
    accept_multiple_files=True, 
    type=["pdf", "jpg", "jpeg", "png"]
)

st.divider()
st.subheader("📧 Configuração de Disparo")
email_destino = st.text_input("E-mail para receber os documentos (Aprovador/Financeiro):", placeholder="financeiro@viaappia.com.br")

# --- LÓGICA DE GERAÇÃO E ENVIO ---
if st.button("Gerar Relatório e Enviar", type="primary"):
    if not os.path.exists(TEMPLATE_PATH):
        st.error(f"Arquivo modelo '{TEMPLATE_PATH}' não encontrado na pasta raiz.")
        st.stop()
        
    with st.spinner('Processando os dados: Preenchendo Excel, convertendo PDF e organizando comprovantes...'):
        try:
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
            ws = wb['FORMULARIO_FI']
            
            # 1. Preencher Cabeçalho
            ws['I5'] = solicitante
            ws['I6'] = solicitante_email
            ws['G10'] = colaborador_nome
            ws['G11'] = colaborador_cpf
            ws['S10'] = nivel_hierarquico
            ws['S11'] = fornecedor_sap
            
            # 2. Preencher Despesas 
            linha_atual = 15
            for index, row in despesas_editadas.iterrows():
                if linha_atual > 34: break
                ws[f'B{linha_atual}'] = row.get("Data (DD/MM/AAA)", "")
                ws[f'D{linha_atual}'] = row.get("Conta Razão", "")
                ws[f'H{linha_atual}'] = row.get("Centro de Custo", "")
                ws[f'L{linha_atual}'] = row.get("Motivo ou Justificativa", "")
                ws[f'S{linha_atual}'] = row.get("Qtde", "")
                ws[f'T{linha_atual}'] = row.get("Valor Gasto (R$)", "")
                linha_atual += 1
                
            # 3. Preencher KM 
            linha_atual = 39
            for index, row in km_editados.iterrows():
                if linha_atual > 46: break
                ws[f'B{linha_atual}'] = row.get("Data (DD/MM/AAA)", "")
                ws[f'D{linha_atual}'] = row.get("Conta Razão", "")
                ws[f'H{linha_atual}'] = row.get("Centro de Custo", "")
                ws[f'L{linha_atual}'] = row.get("Motivo/Origem>Destino", "")
                ws[f'R{linha_atual}'] = row.get("Km (Qtde)", "")
                ws[f'S{linha_atual}'] = row.get("Valor/Km (R$)", "")
                ws[f'T{linha_atual}'] = row.get("Valor Gasto (R$)", "")
                linha_atual += 1
                
            # Salvar Excel temporário
            nome_base = colaborador_nome.replace(" ", "_").strip() if colaborador_nome else "Desconhecido"
            output_xlsx = f"Reembolso_{nome_base}.xlsx"
            output_pdf = output_xlsx.replace(".xlsx", ".pdf")
            
            wb.save(output_xlsx)
            
            # Tenta converter o Excel preenchido para PDF Inicial
            pdf_gerado = False
            try:
                subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf", output_xlsx], check=True)
                if os.path.exists(output_pdf):
                    pdf_gerado = True
            except Exception:
                st.warning("Gerador conversor para PDF falhou. O Excel será enviado no lugar do PDF principal.")

            # --- SESSÃO DE ANEXOS PARA O E-MAIL ---
            anexos_para_email = []
            
            # Anexa o formulário gerado (priorizando PDF, caso falhe, usa o Excel)
            if pdf_gerado:
                anexos_para_email.append(output_pdf)
            else:
                anexos_para_email.append(output_xlsx)
            
            # Salva temporariamente os comprovantes na raiz para criar os arquivos do anexo do email
            for uf in comprovantes_upload:
                temp_path = f"tmp_{uf.name}"
                with open(temp_path, "wb") as f_tmp:
                    f_tmp.write(uf.getbuffer())
                anexos_para_email.append(temp_path)

            # --- SESSÃO DE DISPARO DE E-MAIL ---
            if email_destino:
                corpo_email_html = f"""
                <html>
                <body style="font-family: Arial, sans-serif; color: #111111; background-color: #FFFFFF;">
                    <div style="max-width: 600px; margin: auto; background: white; padding: 20px; border-radius: 8px; border: 1px solid #e1e4e8;">
                        <h2 style="color: #1A5D5C; border-bottom: 2px solid #1A5D5C; padding-bottom: 5px;">Relatório de Reembolsos</h2>
                        <p>Olá,</p>
                        <p>Segue em anexo o formulário de reembolso e os respectivos comprovantes enviados pelo colaborador <b>{colaborador_nome}</b>.</p>
                        <br>
                        <p style="background-color: #F8F9FA; padding: 15px; border-left: 4px solid #1A5D5C; border-radius: 4px;">
                            <b style="color: #1A5D5C;">Solicitante responsável:</b> {solicitante}<br>
                            <span style="font-size: 0.9em; display: inline-block; margin-top: 5px;">Qualquer dúvida, responder diretamente a este e-mail para falar com {solicitante}.</span>
                        </p>
                        <br>
                        <hr style="border: 0; height: 1px; background-color: #E1E4E8;">
                        <p style="font-size: 0.8em; color: gray; text-align: center;">
                            <i>Este é um envio automático do Portal de Reembolsos - Via Appia.</i>
                        </p>
                    </div>
                </body>
                </html>
                """
                
                try:
                    lista_emails = [email_destino.strip()]
                    if solicitante_email:
                        lista_emails.append(solicitante_email.strip())

                    enviar_email_corporativo(
                        remetente_email="naoresponder@viaappia.com.br",
                        destinatarios=lista_emails,
                        assunto=f"Relatório de Reembolso e Comprovantes - {colaborador_nome}",
                        corpo_html=corpo_email_html,
                        reply_to=solicitante_email.strip(),
                        anexos=anexos_para_email
                    )
                    st.success(f"✅ Sucesso! O formulário e {len(comprovantes_upload)} comprovantes foram formatados e enviados para o Financeiro!")
                except Exception as e:
                    st.error(f"Erro ao disparar o e-mail: {e}")
            else:
                st.warning("Nenhum e-mail de destino do financeiro foi preenchido.")

            # Botões para download local
            if pdf_gerado and os.path.exists(output_pdf):
                with open(output_pdf, "rb") as file_pdf:
                    st.download_button(label="📥 Baixar Formulário Oficial (PDF)", data=file_pdf, file_name=output_pdf, mime="application/pdf", type="primary")
            
            with open(output_xlsx, "rb") as file_xlsx:
                st.download_button(label="📥 Baixar apenas a Tabela Excel", data=file_xlsx, file_name=output_xlsx, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    
        except Exception as e:
            st.error(f"Ocorreu um erro geral de processamento: {e}")
